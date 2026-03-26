# pip install fastapi paho-mqtt uvicorn[standard]

import sys
import os
import csv
import signal
import json
import asyncio
import random
import shutil
import time
import threading
import sqlite3
from datetime import datetime, timedelta, timezone, date
from pathlib import Path
from typing import Dict, Any

try:
    import yaml
except ImportError:
    yaml = None  # type: ignore

from fastapi import FastAPI, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
import paho.mqtt.client as mqtt
import uvicorn

from contextlib import asynccontextmanager

BASE_DIR = Path(__file__).resolve().parent
PAGES_DIR = BASE_DIR / "pages"
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = Path("data")
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = str(DATA_DIR / "spots.db")


def _load_config() -> dict:
    cfg: dict = {}
    cfg_path = BASE_DIR / "config.yaml"
    override_path = DATA_DIR / "config.yaml"
    if override_path.exists():
        cfg_path = override_path
    if cfg_path.exists() and yaml is not None:
        with open(cfg_path) as fh:
            cfg = yaml.safe_load(fh) or {}
    max_slots = int((cfg.get("mydx") or {}).get("max_slots", 10))
    env = os.environ.get("MYDX_MAX_SLOTS")
    if env is not None:
        max_slots = int(env)
    max_hours = float((cfg.get("mydx") or {}).get("max_hours", 2))
    env = os.environ.get("MYDX_MAX_HOURS")
    if env is not None:
        max_hours = float(env)
    return {"mydx_max_slots": max_slots, "mydx_max_hours": max_hours}


CONFIG: dict = {}


@asynccontextmanager
async def lifespan(app: FastAPI):
    global main_loop, mqtt_client

    main_loop = asyncio.get_running_loop()
    CONFIG.update(_load_config())
    db_init()
    import_dxpeditions_from_data_dir()
    sync_dxpedition_subscriptions()  # loads active callsigns; MQTT not yet connected

    def _on_sighup():
        print("[dxpedition] SIGHUP received, reimporting...")
        import_dxpeditions_from_data_dir()
        sync_dxpedition_subscriptions()

    main_loop.add_signal_handler(signal.SIGHUP, _on_sighup)

    # Restore MQTT subscriptions for callsigns that were active at last shutdown.
    # Give each a 60s grace period; if no client reconnects in time, unsubscribe.
    open_callsigns = mydx_get_open_callsigns()
    mydx_close_all_sessions()
    for cs in open_callsigns:
        mydx_slots[cs] = {"clients": {}, "release_task": None}
        # MQTT will be subscribed in on_connect via the mydx_slots keys

    mqtt_client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2)
    mqtt_client.on_connect = on_connect
    mqtt_client.on_message = on_message
    mqtt_client.connect(BROKER, PORT)
    mqtt_client.loop_start()

    # Start 60s grace tasks for restored slots (give clients time to reconnect)
    async def _restore_grace(cs: str):
        await asyncio.sleep(60)
        s = mydx_slots.get(cs)
        if s and not s["clients"]:
            del mydx_slots[cs]
            _mydx_unsubscribe(cs)
            print(f"[mydx] grace expired after restart, unsubscribed {cs}")

    for cs in open_callsigns:
        task = asyncio.create_task(_restore_grace(cs))
        mydx_slots[cs]["release_task"] = task

    hb_task = asyncio.create_task(heartbeat_task())
    sync_task = asyncio.create_task(daily_sync_task())
    time_limit_task = asyncio.create_task(mydx_time_limit_task())

    yield

    hb_task.cancel()
    sync_task.cancel()
    time_limit_task.cancel()
    main_loop.remove_signal_handler(signal.SIGHUP)

    if mqtt_client is not None:
        mqtt_client.loop_stop()
        mqtt_client.disconnect()

    if _db is not None:
        with _db_lock:
            _db.close()


app = FastAPI(lifespan=lifespan)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# ==========================
# MQTT設定（PSKReporter）
# ==========================
BROKER = "mqtt.pskreporter.info"
PORT = 1883

TOPIC_FROM_JP = "pskr/filter/v2/+/FT8/+/+/+/+/339/+"
TOPIC_TO_JP = "pskr/filter/v2/+/FT8/+/+/+/+/+/339"

# ==========================
# SQLite（直近15分保持）
# ==========================
KEEP_SEC = 900

_db: sqlite3.Connection | None = None
_db_lock = threading.Lock()

clients: Dict[WebSocket, Dict[str, Any]] = {}
main_loop: asyncio.AbstractEventLoop | None = None
mqtt_client: mqtt.Client | None = None

HB_INTERVAL = 10
last_mqtt_ts_from_jp = 0.0
last_mqtt_ts_to_jp = 0.0
last_mqtt_ts_dxpedition = 0.0

dxpedition_subscribed_callsigns: set[str] = set()

MYDX_KEEP_SEC = 900  # 15 minutes

# mycall proxy slots:
#   { callsign: {"clients": {ws: txrx}, "release_task": Task|None} }
mydx_slots: dict[str, dict] = {}


def get_mydx_max_slots() -> int:
    return CONFIG.get("mydx_max_slots", 10)


def get_mydx_max_seconds() -> float:
    return CONFIG.get("mydx_max_hours", 2) * 3600


def maidenhead_to_latlon(locator: str):
    locator = locator.strip().upper()
    lon = (ord(locator[0]) - ord('A')) * 20 - 180
    lat = (ord(locator[1]) - ord('A')) * 10 - 90
    lon += int(locator[2]) * 2
    lat += int(locator[3]) * 1
    lon_size = 2
    lat_size = 1
    if len(locator) >= 6:
        lon += (ord(locator[4]) - ord('A')) * (5 / 60)
        lat += (ord(locator[5]) - ord('A')) * (2.5 / 60)
        lon_size = 5 / 60
        lat_size = 2.5 / 60
    lon += lon_size / 2
    lat += lat_size / 2
    return lat, lon


def apply_blur(lat, lon, locator: str):
    locator = locator.strip().upper()
    if len(locator) == 4:
        lat += random.uniform(-0.5, 0.5)
        lon += random.uniform(-1.0, 1.0)
    return lat, lon


def db_init():
    global _db
    _db = sqlite3.connect(DB_PATH, check_same_thread=False)
    _db.row_factory = sqlite3.Row
    with _db_lock:
        _db.execute(
            """
            CREATE TABLE IF NOT EXISTS spots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts REAL NOT NULL,
                payload TEXT NOT NULL
            )
            """
        )
        _db.execute("CREATE INDEX IF NOT EXISTS idx_spots_ts ON spots(ts)")
        _db.execute(
            """
            CREATE TABLE IF NOT EXISTS dxpedition (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                callsign    TEXT    NOT NULL,
                entity_name TEXT,
                dxcc        INTEGER,
                grid        TEXT,
                start_dt    TEXT,
                end_dt      TEXT,
                url         TEXT,
                notes       TEXT
            )
            """
        )
        _db.execute(
            "CREATE INDEX IF NOT EXISTS idx_dxpedition_callsign ON dxpedition(callsign)")
        _db.execute(
            "CREATE INDEX IF NOT EXISTS idx_dxpedition_dxcc ON dxpedition(dxcc)")
        _db.execute(
            "CREATE INDEX IF NOT EXISTS idx_dxpedition_dates ON dxpedition(start_dt, end_dt)")
        _db.execute(
            """
            CREATE TABLE IF NOT EXISTS mydx_spots (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                ts      REAL    NOT NULL,
                mycall  TEXT    NOT NULL,
                txrx    TEXT    NOT NULL,
                payload TEXT    NOT NULL
            )
            """
        )
        _db.execute(
            "CREATE INDEX IF NOT EXISTS idx_mydx_spots_mycall_ts ON mydx_spots(mycall, ts)")
        _db.execute(
            """
            CREATE TABLE IF NOT EXISTS mydx_sessions (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                mycall         TEXT    NOT NULL,
                connected_at   REAL    NOT NULL,
                disconnected_at REAL
            )
            """
        )
        _db.execute(
            "CREATE INDEX IF NOT EXISTS idx_mydx_sessions_mycall ON mydx_sessions(mycall)")
        _db.execute(
            """
            CREATE TABLE IF NOT EXISTS dxpedition_activity (
                callsign   TEXT    NOT NULL,
                hour_utc   TEXT    NOT NULL,
                band       INTEGER NOT NULL,
                spot_count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (callsign, hour_utc, band)
            )
            """
        )
        _db.execute(
            "CREATE INDEX IF NOT EXISTS idx_dxped_activity ON dxpedition_activity(callsign, hour_utc)")
        _db.commit()


DXPEDITION_IMPORT_COLUMNS = {
    "callsign", "entity_name", "dxcc", "grid",
    "start_dt", "end_dt", "url", "notes",
}

DXPEDITION_BACKUP_DIR = DATA_DIR / "backup"


def _normalize_callsign_field(value: str) -> str:
    """Normalize a comma-separated callsign field: strip, uppercase, rejoin."""
    parts = [p.strip().upper() for p in value.split(",") if p.strip()]
    return ",".join(parts)


def _normalize_dxpedition_value(key: str, value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        if key == "callsign":
            return _normalize_callsign_field(value)
        return value
    return value


def _load_dxpedition_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"[dxpedition] openpyxl not installed, skipping {path}", file=sys.stderr)
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        all_empty = True
        for key, value in zip(header, row):
            if not key or key not in DXPEDITION_IMPORT_COLUMNS:
                continue
            norm = _normalize_dxpedition_value(key, value)
            record[key] = norm
            if norm is not None:
                all_empty = False
        if not all_empty:
            records.append(record)
    return records


def _load_dxpedition_csv(path: Path) -> list[dict]:
    records = []
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh, delimiter=",")
        for row in reader:
            record = {}
            all_empty = True
            for key, value in row.items():
                key = (key or "").strip()
                if not key or key not in DXPEDITION_IMPORT_COLUMNS:
                    continue
                norm = _normalize_dxpedition_value(key, value)
                record[key] = norm
                if norm is not None:
                    all_empty = False
            if not all_empty:
                records.append(record)
    return records


def _insert_dxpedition_record(db: sqlite3.Connection, record: dict):
    callsign = (record.get("callsign") or "").strip()
    if not callsign:
        return
    db.execute(
        """
        INSERT INTO dxpedition(callsign, entity_name, dxcc, grid, start_dt, end_dt, url, notes)
        VALUES(?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (callsign, record.get("entity_name"), record.get("dxcc"),
         record.get("grid"), record.get("start_dt"), record.get("end_dt"),
         record.get("url"), record.get("notes")),
    )


def _move_to_backup(src: Path) -> Path:
    DXPEDITION_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    dst = DXPEDITION_BACKUP_DIR / src.name
    if dst.exists():
        stem, suffix = src.stem, src.suffix
        counter = 1
        while True:
            candidate = DXPEDITION_BACKUP_DIR / f"{stem}_{counter}{suffix}"
            if not candidate.exists():
                dst = candidate
                break
            counter += 1
    shutil.move(str(src), str(dst))
    return dst


def import_dxpeditions_from_data_dir():
    """Replace dxpedition table from any DX*.xlsx / DX*.tsv files in DATA_DIR."""
    assert _db is not None
    files = sorted(DATA_DIR.glob("DX*.xlsx")) + sorted(DATA_DIR.glob("DX*.csv"))
    if not files:
        return

    print(f"[dxpedition] found {len(files)} import file(s), replacing table")
    with _db_lock:
        _db.execute("DELETE FROM dxpedition")
        total = 0
        for path in files:
            try:
                if path.suffix.lower() == ".xlsx":
                    records = _load_dxpedition_xlsx(path)
                else:
                    records = _load_dxpedition_csv(path)
                for record in records:
                    _insert_dxpedition_record(_db, record)
                _db.commit()
                backup = _move_to_backup(path)
                print(f"[dxpedition] imported {len(records)} record(s) from {path.name} → {backup}")
                total += len(records)
            except Exception as e:
                _db.rollback()
                print(f"[dxpedition] error importing {path}: {e}", file=sys.stderr)
    print(f"[dxpedition] import complete: {total} total record(s)")


def get_active_dxpeditions() -> list[dict]:
    assert _db is not None
    with _db_lock:
        cur = _db.execute(
            """
            SELECT * FROM dxpedition
            WHERE (start_dt IS NULL OR start_dt <= date('now'))
              AND (end_dt   IS NULL OR end_dt   >= date('now'))
            ORDER BY callsign ASC
            """
        )
        return [dict(row) for row in cur.fetchall()]


def sync_dxpedition_subscriptions():
    global dxpedition_subscribed_callsigns
    active: set[str] = set()
    for row in get_active_dxpeditions():
        for cs in row["callsign"].split(","):
            cs = cs.strip().upper()
            if cs:
                active.add(cs)
    if mqtt_client is not None:
        for cs in dxpedition_subscribed_callsigns - active:
            mqtt_client.unsubscribe(f"pskr/filter/v2/+/+/{cs.replace('/', '.')}/#")
            print("Unsubscribed: %s", f"pskr/filter/v2/+/+/{cs}/#")
        for cs in active - dxpedition_subscribed_callsigns:
            mqtt_client.subscribe(f"pskr/filter/v2/+/+/{cs.replace('/', '.')}/#")
            print("Subscribed: %s", f"pskr/filter/v2/+/+/{cs}/#")
    dxpedition_subscribed_callsigns = active


def db_upsert_activity(callsign: str, band: int, hour_utc: str):
    assert _db is not None
    with _db_lock:
        _db.execute(
            """
            INSERT INTO dxpedition_activity (callsign, hour_utc, band, spot_count)
            VALUES (?, ?, ?, 1)
            ON CONFLICT (callsign, hour_utc, band)
            DO UPDATE SET spot_count = spot_count + 1
            """,
            (callsign, hour_utc, band),
        )
        _db.commit()


def db_insert(payload: str):
    assert _db is not None
    now = time.time()
    cutoff = now - KEEP_SEC
    with _db_lock:
        _db.execute("INSERT INTO spots(ts, payload) VALUES(?, ?)",
                    (now, payload))
        _db.execute("DELETE FROM spots WHERE ts < ?", (cutoff,))
        _db.commit()


def db_select_recent(mode: str | None = None, dxcall: str | None = None, keep_sec: int = KEEP_SEC) -> list[str]:
    assert _db is not None
    now = time.time()
    cutoff = now - keep_sec
    with _db_lock:
        cur = _db.execute(
            "SELECT payload FROM spots WHERE ts >= ? ORDER BY ts ASC",
            (cutoff,),
        )
        rows = [row[0] for row in cur.fetchall()]

    if mode is None:
        return rows

    filtered: list[str] = []
    for payload in rows:
        try:
            obj = json.loads(payload)
            if obj.get("type") == "spot" and obj.get("mode") == mode:
                if mode == "dxpedition" and dxcall and obj.get("dxcall", "") != dxcall:
                    continue
                filtered.append(payload)
        except Exception:
            pass
    return filtered


def mydx_db_insert(mycall: str, ts: float, txrx: str, payload: str):
    assert _db is not None
    cutoff = ts - MYDX_KEEP_SEC
    with _db_lock:
        _db.execute(
            "INSERT INTO mydx_spots(ts, mycall, txrx, payload) VALUES(?,?,?,?)",
            (ts, mycall, txrx, payload),
        )
        _db.execute(
            "DELETE FROM mydx_spots WHERE mycall=? AND ts<?",
            (mycall, cutoff),
        )
        _db.commit()


def mydx_db_select_recent(mycall: str, txrx: str) -> list[str]:
    assert _db is not None
    cutoff = time.time() - MYDX_KEEP_SEC
    with _db_lock:
        cur = _db.execute(
            "SELECT payload FROM mydx_spots WHERE mycall=? AND txrx=? AND ts>=? ORDER BY ts ASC",
            (mycall, txrx, cutoff),
        )
        return [row[0] for row in cur.fetchall()]


def mydx_session_open(mycall: str) -> int:
    assert _db is not None
    now = time.time()
    with _db_lock:
        cur = _db.execute(
            "INSERT INTO mydx_sessions(mycall, connected_at) VALUES(?, ?)",
            (mycall, now),
        )
        _db.commit()
        return cur.lastrowid  # type: ignore[return-value]


def mydx_session_close(session_id: int):
    assert _db is not None
    with _db_lock:
        _db.execute(
            "UPDATE mydx_sessions SET disconnected_at=? WHERE id=?",
            (time.time(), session_id),
        )
        _db.commit()


def mydx_get_open_callsigns() -> list[str]:
    """Return callsigns with sessions that were open (no disconnected_at) at last shutdown."""
    assert _db is not None
    with _db_lock:
        cur = _db.execute(
            "SELECT DISTINCT mycall FROM mydx_sessions WHERE disconnected_at IS NULL"
        )
        return [row[0] for row in cur.fetchall()]


def mydx_used_seconds(mycall: str) -> float:
    """Return total connected seconds for mycall in the last 24 hours.
    Overlapping sessions (e.g. display WS + background keepalive WS open simultaneously)
    are merged so they are not double-counted."""
    assert _db is not None
    since = time.time() - 86400
    with _db_lock:
        cur = _db.execute(
            """
            SELECT connected_at, disconnected_at FROM mydx_sessions
            WHERE mycall=? AND (disconnected_at IS NULL OR disconnected_at >= ?)
            """,
            (mycall, since),
        )
        now = time.time()
        intervals = []
        for connected_at, disconnected_at in cur.fetchall():
            start = max(connected_at, since)
            end = disconnected_at if disconnected_at is not None else now
            if end > start:
                intervals.append((start, end))
        if not intervals:
            return 0.0
        intervals.sort()
        total = 0.0
        cur_start, cur_end = intervals[0]
        for start, end in intervals[1:]:
            if start <= cur_end:
                cur_end = max(cur_end, end)
            else:
                total += cur_end - cur_start
                cur_start, cur_end = start, end
        total += cur_end - cur_start
        return total


def mydx_close_all_sessions():
    """Mark all open sessions as closed (called at startup to clean up stale records)."""
    assert _db is not None
    now = time.time()
    with _db_lock:
        _db.execute(
            "UPDATE mydx_sessions SET disconnected_at=? WHERE disconnected_at IS NULL",
            (now,),
        )
        _db.commit()


def mode_from_topic(topic: str) -> str | None:
    parts = topic.split("/")
    if len(parts) >= 2 and parts[-2] == "339":
        return "from_jp"
    if len(parts) >= 1 and parts[-1] == "339":
        return "to_jp"
    return None


def should_forward_spot(mode: str, ra: int | None, sa: int | None) -> bool:
    if mode == "from_jp":
        return ra != 339
    if mode == "to_jp":
        return sa != 339
    return True


def should_forward_local_spot(mode: str, ra: int | None, sa: int | None) -> bool:
    if mode == "from_jp":
        return ra == 339
    if mode == "to_jp":
        return sa == 339
    return False


def _norm_call(value: str | None) -> str:
    return (value or "").strip().upper()


def should_forward_mydx_spot(mode: str, sc: str | None, rc: str | None, mycall: str) -> bool:
    mycall = _norm_call(mycall)
    if not mycall:
        return True

    sc = _norm_call(sc)
    rc = _norm_call(rc)

    if mode == "from_jp":
        return sc == mycall
    if mode == "to_jp":
        return rc == mycall
    return False


async def broadcast(message: str):
    dead: list[WebSocket] = []
    is_hb = False
    msg_mode: str | None = None
    obj: dict[str, Any] = {}
    try:
        obj = json.loads(message)
        if obj.get("type") == "hb":
            is_hb = True
        elif obj.get("type") == "spot":
            msg_mode = obj.get("mode")
    except Exception:
        pass

    for ws, info in list(clients.items()):
        if not info.get("ready", False):
            continue
        try:
            if is_hb:
                await ws.send_text(message)
            else:
                if msg_mode is None:
                    continue
                if info.get("mode") != msg_mode:
                    continue
                if msg_mode == "dxpedition":
                    client_dxcall = info.get("dxcall", "")
                    if client_dxcall and client_dxcall != obj.get("dxcall", ""):
                        continue
                elif info.get("local"):
                    if not should_forward_local_spot(msg_mode, obj.get("ra"), obj.get("sa")):
                        continue
                else:
                    if not should_forward_spot(msg_mode, obj.get("ra"), obj.get("sa")):
                        continue
                mycall = info.get("mycall", "")
                if mycall and not should_forward_mydx_spot(msg_mode, obj.get("sc"), obj.get("rc"), mycall):
                    continue
                await ws.send_text(message)
        except Exception:
            dead.append(ws)

    for ws in dead:
        clients.pop(ws, None)


async def daily_sync_task():
    while True:
        now = datetime.now(timezone.utc)
        next_midnight = (now + timedelta(days=1)).replace(hour=0,
                                                          minute=0, second=0, microsecond=0)
        await asyncio.sleep((next_midnight - now).total_seconds())
        sync_dxpedition_subscriptions()


async def mydx_time_limit_task():
    while True:
        await asyncio.sleep(600)  # check every 10 minutes
        now = time.time()
        expired = [
            cs for cs, slot in list(mydx_slots.items())
            if slot.get("expires_at") is not None and now >= slot["expires_at"]
        ]
        for cs in expired:
            slot = mydx_slots.get(cs)
            if not slot:
                continue
            print(f"[mydx] time limit exceeded, closing slot {cs}")
            msg = json.dumps({"type": "time_limit_exceeded"})
            for ws in list(slot["clients"]):
                try:
                    await ws.send_text(msg)
                    await ws.close()
                except Exception:
                    pass


async def heartbeat_task():
    while True:
        hb = json.dumps(
            {
                "type": "hb",
                "ts": time.time(),
                "last_mqtt_ts_from_jp": last_mqtt_ts_from_jp,
                "last_mqtt_ts_to_jp": last_mqtt_ts_to_jp,
                "last_mqtt_ts_dxpedition": last_mqtt_ts_dxpedition,
            }
        )
        await broadcast(hb)
        await asyncio.sleep(HB_INTERVAL)


def _mydx_subscribe(mycall: str):
    if mqtt_client is None:
        return
    enc = mycall.replace("/", ".")
    mqtt_client.subscribe(f"pskr/filter/v2/+/+/{enc}/#")
    mqtt_client.subscribe(f"pskr/filter/v2/+/+/+/{enc}/#")
    print(f"[mydx] subscribed {mycall}, slots={len(mydx_slots)}/{get_mydx_max_slots()}")


def _mydx_unsubscribe(mycall: str):
    if mqtt_client is None:
        return
    enc = mycall.replace("/", ".")
    mqtt_client.unsubscribe(f"pskr/filter/v2/+/+/{enc}/#")
    mqtt_client.unsubscribe(f"pskr/filter/v2/+/+/+/{enc}/#")
    print(f"[mydx] unsubscribed {mycall}, slots={len(mydx_slots)}/{get_mydx_max_slots()}")


async def _mydx_broadcast_slots():
    """Push updated slot count to all active mydx clients."""
    max_slots = get_mydx_max_slots()
    msg = json.dumps({"type": "slots", "used": len(mydx_slots), "max": max_slots})
    for slot in list(mydx_slots.values()):
        for ws in list(slot["clients"]):
            try:
                await ws.send_text(msg)
            except Exception:
                pass


async def _mydx_dispatch(mycall: str, role: str, data: dict):
    """Compute lat/lon and forward a mydx spot to all matching clients."""
    slot = mydx_slots.get(mycall)
    if not slot or not slot["clients"]:
        return

    rl = data.get("rl")
    sl = data.get("sl")
    sc = data.get("sc", "")
    rc = data.get("rc", "")
    b = data.get("b", "")
    ts = time.time()

    rl_lat = rl_lon = None
    sl_lat = sl_lon = None
    if rl:
        try:
            rl_lat, rl_lon = maidenhead_to_latlon(rl)
            rl_lat, rl_lon = apply_blur(rl_lat, rl_lon, rl)
        except Exception:
            pass
    if sl:
        try:
            sl_lat, sl_lon = maidenhead_to_latlon(sl)
            sl_lat, sl_lon = apply_blur(sl_lat, sl_lon, sl)
        except Exception:
            pass

    snr_val = data.get("rp")
    freq_val = data.get("f")
    spot_mode_val = data.get("spot_mode", "")

    payloads: dict[str, str] = {}
    if role == "sc" and rl_lat is not None:
        payloads["tx"] = json.dumps({
            "type": "spot", "mode": "mydx", "txrx": "tx",
            "lat": rl_lat, "lon": rl_lon,
            "b": b, "sc": sc, "rc": rc, "ts": ts,
            "snr": snr_val, "f": freq_val, "spot_mode": spot_mode_val,
        })
        mydx_db_insert(mycall, ts, "tx", payloads["tx"])
    if role == "rc" and sl_lat is not None:
        payloads["rx"] = json.dumps({
            "type": "spot", "mode": "mydx", "txrx": "rx",
            "lat": sl_lat, "lon": sl_lon,
            "b": b, "sc": sc, "rc": rc, "ts": ts,
            "snr": snr_val, "f": freq_val, "spot_mode": spot_mode_val,
        })
        mydx_db_insert(mycall, ts, "rx", payloads["rx"])

    for ws, txrx in list(slot["clients"].items()):
        payload = payloads.get(txrx)
        if payload:
            try:
                await ws.send_text(payload)
            except Exception:
                pass


def on_connect(client, userdata, flags, reason_code, properties):
    print("Connected: %s", reason_code)
    client.subscribe(TOPIC_FROM_JP)
    client.subscribe(TOPIC_TO_JP)
    for cs in dxpedition_subscribed_callsigns:
        client.subscribe(f"pskr/filter/v2/+/+/{cs.replace('/', '.')}/#")
        print("Subscribed: %s", f"pskr/filter/v2/+/+/{cs}/#")
    for cs in mydx_slots:
        enc = cs.replace("/", ".")
        client.subscribe(f"pskr/filter/v2/+/+/{enc}/#")
        client.subscribe(f"pskr/filter/v2/+/+/+/{enc}/#")
    print("Subscribed: %s", TOPIC_FROM_JP)
    print("Subscribed: %s", TOPIC_TO_JP)


def on_message(client, userdata, msg):
    global last_mqtt_ts_from_jp, last_mqtt_ts_to_jp, last_mqtt_ts_dxpedition
    try:
        mode = mode_from_topic(msg.topic)
        has_mydx = bool(mydx_slots)

        if mode is None and not has_mydx and not dxpedition_subscribed_callsigns:
            return

        data = json.loads(msg.payload.decode())
        rl = data.get("rl")
        sl = data.get("sl")
        sc = data.get("sc")
        rc = data.get("rc")
        sc_upper = (sc or "").upper()
        topic_parts = msg.topic.split("/")
        spot_mode = topic_parts[4] if len(topic_parts) > 4 else ""
        snr = data.get("rp")   # received power / signal report (dB)
        freq = data.get("f")

        if mode is None and sc_upper in dxpedition_subscribed_callsigns:
            mode = "dxpedition"

        # --- from_jp / to_jp / dxpedition processing ---
        if mode is not None:
            rl_lat = rl_lon = None
            if rl:
                rl_lat, rl_lon = maidenhead_to_latlon(rl)
                rl_lat, rl_lon = apply_blur(rl_lat, rl_lon, rl)

            sl_lat = sl_lon = None
            if sl:
                try:
                    sl_lat, sl_lon = maidenhead_to_latlon(sl)
                    sl_lat, sl_lon = apply_blur(sl_lat, sl_lon, sl)
                except Exception:
                    pass

            dxcall = None
            send_mode = True
            if mode == "from_jp":
                if rl_lat is None or rl_lon is None:
                    send_mode = False
                else:
                    marker_lat, marker_lon = rl_lat, rl_lon
                    peer_lat, peer_lon = sl_lat, sl_lon
                    last_mqtt_ts_from_jp = time.time()
            elif mode == "to_jp":
                if sl_lat is None or sl_lon is None:
                    send_mode = False
                else:
                    marker_lat, marker_lon = sl_lat, sl_lon
                    peer_lat, peer_lon = rl_lat, rl_lon
                    last_mqtt_ts_to_jp = time.time()
            elif mode == "dxpedition":
                if rl_lat is None or rl_lon is None:
                    send_mode = False
                else:
                    marker_lat, marker_lon = rl_lat, rl_lon
                    peer_lat, peer_lon = sl_lat, sl_lon
                    dxcall = sc_upper or None
                    last_mqtt_ts_dxpedition = time.time()
            else:
                send_mode = False

            if send_mode:
                send_data = json.dumps(
                    {
                        "type": "spot",
                        "mode": mode,
                        "lat": marker_lat,
                        "lon": marker_lon,
                        "peer_lat": peer_lat,
                        "peer_lon": peer_lon,
                        "b": data.get("b", ""),
                        "ra": data.get("ra"),
                        "sa": data.get("sa"),
                        "sc": sc,
                        "rc": rc,
                        "dxcall": dxcall,
                        "snr": snr,
                        "f": freq,
                        "spot_mode": spot_mode,
                        "ts": time.time(),
                    }
                )
                db_insert(send_data)
                if mode == "dxpedition" and dxcall:
                    try:
                        band_int = int((data.get("b") or "0").replace("m", "") or "0")
                        hour_utc = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H")
                        db_upsert_activity(dxcall, band_int, hour_utc)
                    except Exception:
                        pass
                if main_loop is not None:
                    main_loop.call_soon_threadsafe(
                        asyncio.create_task, broadcast(send_data))

        # --- mydx proxy routing ---
        if has_mydx:
            rc_upper = (rc or "").upper()
            data_aug = {**data, "spot_mode": spot_mode}
            if sc_upper in mydx_slots:
                if main_loop is not None:
                    main_loop.call_soon_threadsafe(
                        asyncio.create_task, _mydx_dispatch(sc_upper, "sc", data_aug))
            elif rc_upper in mydx_slots:
                if main_loop is not None:
                    main_loop.call_soon_threadsafe(
                        asyncio.create_task, _mydx_dispatch(rc_upper, "rc", data_aug))

    except Exception as exc:
        print("Error: %s", exc)


async def _handle_mydx_ws(websocket: WebSocket):
    mycall = websocket.query_params.get("mycall", "").strip().upper()
    txrx = websocket.query_params.get("txrx", "tx")
    if txrx not in ("tx", "rx"):
        txrx = "tx"

    if not mycall:
        await websocket.close(1008)
        return

    max_slots = get_mydx_max_slots()
    max_seconds = get_mydx_max_seconds()

    # Check slot availability (0 = unlimited)
    if max_slots > 0 and mycall not in mydx_slots and len(mydx_slots) >= max_slots:
        await websocket.send_text(json.dumps({
            "type": "unavailable",
            "used": len(mydx_slots),
            "max": max_slots,
        }))
        await websocket.close()
        return

    # Check time limit (0 = unlimited)
    if max_seconds > 0 and mycall not in mydx_slots:
        used = mydx_used_seconds(mycall)
        remaining = max_seconds - used
        if remaining <= 0:
            await websocket.send_text(json.dumps({"type": "time_limit_exceeded"}))
            await websocket.close()
            return
        expires_at = time.time() + remaining
    else:
        expires_at = None

    # Acquire or reuse slot
    if mycall not in mydx_slots:
        mydx_slots[mycall] = {"clients": {}, "release_task": None, "expires_at": expires_at}
        _mydx_subscribe(mycall)
        await _mydx_broadcast_slots()
    else:
        slot = mydx_slots[mycall]
        if slot["release_task"] and not slot["release_task"].done():
            slot["release_task"].cancel()
            slot["release_task"] = None

    slot = mydx_slots[mycall]
    slot["clients"][websocket] = txrx

    session_id = mydx_session_open(mycall)

    # Send current slot status to this client
    await websocket.send_text(json.dumps({
        "type": "slots",
        "used": len(mydx_slots),
        "max": max_slots,
    }))

    # Replay last 15 min from DB
    for payload in mydx_db_select_recent(mycall, txrx):
        await websocket.send_text(payload)

    try:
        while True:
            await websocket.receive_text()
    except (WebSocketDisconnect, Exception):
        pass
    finally:
        mydx_session_close(session_id)
        slot = mydx_slots.get(mycall)
        if slot:
            slot["clients"].pop(websocket, None)
            if not slot["clients"]:
                if slot.get("immediate_release"):
                    del mydx_slots[mycall]
                    _mydx_unsubscribe(mycall)
                    asyncio.create_task(_mydx_broadcast_slots())
                else:
                    async def _grace_release(cs: str = mycall):
                        await asyncio.sleep(30)
                        s = mydx_slots.get(cs)
                        if s and not s["clients"]:
                            del mydx_slots[cs]
                            _mydx_unsubscribe(cs)
                            await _mydx_broadcast_slots()
                    slot["release_task"] = asyncio.create_task(_grace_release())
            else:
                # Other clients remain — the immediate_release was for the disconnecting
                # browser only; clear the flag so remaining browsers get normal grace.
                slot.pop("immediate_release", None)


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()

    mode = websocket.query_params.get("mode", "from_jp")

    if mode == "mydx":
        await _handle_mydx_ws(websocket)
        return

    if mode not in ("from_jp", "to_jp", "dxpedition"):
        mode = "from_jp"

    local = websocket.query_params.get("local") == "1"
    mycall = websocket.query_params.get("mycall", "").strip().upper()
    dxcall = websocket.query_params.get("dxcall", "").strip().upper()
    clients[websocket] = {"ready": False,
                          "mode": mode, "local": local, "mycall": mycall, "dxcall": dxcall}

    try:
        history = db_select_recent(
            mode=mode,
            dxcall=dxcall if mode == "dxpedition" else None,
            keep_sec=180 if mode in ("from_jp", "to_jp") else KEEP_SEC,
        )
        for payload in history:
            try:
                obj = json.loads(payload)
                if obj.get("type") == "spot":
                    if mode != "dxpedition":
                        if local:
                            if not should_forward_local_spot(mode, obj.get("ra"), obj.get("sa")):
                                continue
                        else:
                            if not should_forward_spot(mode, obj.get("ra"), obj.get("sa")):
                                continue
                        mycall_val = clients[websocket].get("mycall", "")
                        if mycall_val and not should_forward_mydx_spot(mode, obj.get("sc"), obj.get("rc"), mycall_val):
                            continue
            except Exception:
                pass
            await websocket.send_text(payload)

        clients[websocket]["ready"] = True

        while True:
            await websocket.receive_text()

    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        clients.pop(websocket, None)


@app.post("/api/mydx_release")
async def api_mydx_release(request: Request):
    """Signal intentional release of a mydx slot — skips the grace timer."""
    data = await request.json()
    mycall = (data.get("mycall") or "").strip().upper()
    if not mycall:
        return {"ok": False}
    slot = mydx_slots.get(mycall)
    if slot:
        slot["immediate_release"] = True
        if not slot["clients"]:
            # WebSocket already closed and grace timer is running — cancel and release now
            rt = slot.get("release_task")
            if rt and not rt.done():
                rt.cancel()
            del mydx_slots[mycall]
            _mydx_unsubscribe(mycall)
            asyncio.create_task(_mydx_broadcast_slots())
    return {"ok": True}


@app.get("/api/dxpeditions")
def api_list_dxpeditions():
    return get_active_dxpeditions()


@app.get("/api/dxpedition_activity")
def api_dxpedition_activity(callsign: str = ""):
    callsigns = [cs.strip() for cs in callsign.split(",") if cs.strip()]
    if not callsigns:
        return []
    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%dT%H")
    assert _db is not None
    with _db_lock:
        placeholders = ",".join("?" * len(callsigns))
        cur = _db.execute(
            f"""
            SELECT callsign, hour_utc, band, spot_count
            FROM dxpedition_activity
            WHERE callsign IN ({placeholders})
              AND hour_utc >= ?
            ORDER BY hour_utc ASC
            """,
            (*callsigns, cutoff),
        )
        return [dict(row) for row in cur.fetchall()]


def page_response(name: str) -> FileResponse:
    return FileResponse(PAGES_DIR / name)


@app.get("/dx")
def index_dx():
    return page_response("app.html")



@app.get("/local")
def index_local():
    return page_response("app.html")


@app.get("/my_dx")
def index_my_dx():
    return page_response("app.html")


@app.get("/dxpedition")
def index_dxpedition():
    return page_response("app.html")


@app.get("/")
def redirect_root():
    return RedirectResponse(url="/dx", status_code=307)


if __name__ == "__main__":
    import logging

    class _SuppressConnectionLogs(logging.Filter):
        def filter(self, record: logging.LogRecord) -> bool:
            msg = record.getMessage()
            return "connection open" not in msg.lower() and "connection closed" not in msg.lower()

    logging.getLogger("uvicorn.error").addFilter(_SuppressConnectionLogs())

    uvicorn.run(app, host="0.0.0.0", port=8000, log_config=None)
