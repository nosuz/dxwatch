#!/usr/bin/env python3
"""Import DX-pedition records from data/DX*.xlsx into SQLite.

Behavior:
- Reads all Excel files matching data/DX*.xlsx
- Uses the first sheet in each workbook
- Treats the first row as header
- Upserts records into data/spots.db
- If 'id' column exists and has a value, update that row by id
- Otherwise upsert by callsign (case-insensitive)
- Moves successfully processed xlsx files into data/backup/

Expected useful columns:
    id, callsign, entity_name, dxcc, grid, start_dt, end_dt, url, notes

Extra columns such as created_at / updated_at are allowed and ignored on import.
"""

import shutil
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

DATA_DIR = Path("data")
BACKUP_DIR = DATA_DIR / "backup"
DB_PATH = DATA_DIR / "spots.db"
FILE_GLOB = "DX*.xlsx"

IMPORT_COLUMNS = {
    "id",
    "callsign",
    "entity_name",
    "dxcc",
    "grid",
    "start_dt",
    "end_dt",
    "url",
    "notes",
    "created_at",
    "updated_at",
}

REQUIRED_IF_NO_ID = "callsign"


def normalize_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def ensure_schema(db: sqlite3.Connection) -> None:
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS dxpedition (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            callsign    TEXT    NOT NULL,
            entity_name TEXT,
            dxcc        INTEGER,
            grid        TEXT,
            start_dt    TEXT,
            end_dt      TEXT,
            url         TEXT,
            notes       TEXT,
            created_at  TEXT NOT NULL DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ','now')),
            updated_at  TEXT NOT NULL DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ','now'))
        )
        """
    )
    db.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_dxpedition_callsign
        ON dxpedition(callsign)
        """
    )
    db.commit()


def load_records_from_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_set = {h for h in header if h}

    if "id" not in header_set and REQUIRED_IF_NO_ID not in header_set:
        raise ValueError("either 'id' or 'callsign' column is required")

    records = []
    for row_index, row in enumerate(rows[1:], start=2):
        record = {}
        is_all_empty = True

        for key, value in zip(header, row):
            if not key:
                continue
            if key not in IMPORT_COLUMNS:
                continue
            norm = normalize_value(value)
            record[key] = norm
            if norm is not None:
                is_all_empty = False

        if is_all_empty:
            continue

        # id があれば整数化
        if record.get("id") is not None:
            try:
                record["id"] = int(record["id"])
            except Exception as e:
                raise ValueError(
                    f"row {row_index}: invalid id: {record['id']}") from e

        # dxcc があれば整数化
        if record.get("dxcc") is not None:
            try:
                record["dxcc"] = int(record["dxcc"])
            except Exception as e:
                raise ValueError(
                    f"row {row_index}: invalid dxcc: {record['dxcc']}") from e

        # id が無い場合は callsign 必須
        if record.get("id") is None:
            callsign = (record.get("callsign") or "").strip()
            if not callsign:
                raise ValueError(
                    f"row {row_index}: callsign is required when id is not specified")

        records.append(record)

    return records


def update_by_id(db: sqlite3.Connection, record: dict) -> str:
    record_id = record.get("id")
    if record_id is None:
        raise ValueError("'id' is required for update_by_id")

    cur = db.execute("SELECT id FROM dxpedition WHERE id = ?", (record_id,))
    row = cur.fetchone()
    if not row:
        raise ValueError(f"id not found: {record_id}")

    callsign = record.get("callsign")
    if callsign is not None:
        callsign = callsign.strip().upper()
        if not callsign:
            raise ValueError("callsign cannot be empty when specified")
    else:
        cur = db.execute(
            "SELECT callsign FROM dxpedition WHERE id = ?", (record_id,))
        existing = cur.fetchone()
        callsign = existing[0]

    db.execute(
        """
        UPDATE dxpedition
        SET callsign=?,
            entity_name=?,
            dxcc=?,
            grid=?,
            start_dt=?,
            end_dt=?,
            url=?,
            notes=?,
            updated_at=strftime('%Y-%m-%dT%H:%M:%SZ','now')
        WHERE id=?
        """,
        (
            callsign,
            record.get("entity_name"),
            record.get("dxcc"),
            record.get("grid"),
            record.get("start_dt"),
            record.get("end_dt"),
            record.get("url"),
            record.get("notes"),
            record_id,
        ),
    )
    return "updated"


def upsert_by_callsign(db: sqlite3.Connection, record: dict) -> str:
    callsign = (record.get("callsign") or "").strip().upper()
    if not callsign:
        raise ValueError("'callsign' field is required")

    cur = db.execute(
        "SELECT id FROM dxpedition WHERE callsign = ?", (callsign,))
    row = cur.fetchone()

    if row:
        db.execute(
            """
            UPDATE dxpedition
            SET entity_name=?,
                dxcc=?,
                grid=?,
                start_dt=?,
                end_dt=?,
                url=?,
                notes=?,
                updated_at=strftime('%Y-%m-%dT%H:%M:%SZ','now')
            WHERE callsign=?
            """,
            (
                record.get("entity_name"),
                record.get("dxcc"),
                record.get("grid"),
                record.get("start_dt"),
                record.get("end_dt"),
                record.get("url"),
                record.get("notes"),
                callsign,
            ),
        )
        return "updated"
    else:
        db.execute(
            """
            INSERT INTO dxpedition(
                callsign, entity_name, dxcc, grid, start_dt, end_dt, url, notes
            )
            VALUES(?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                callsign,
                record.get("entity_name"),
                record.get("dxcc"),
                record.get("grid"),
                record.get("start_dt"),
                record.get("end_dt"),
                record.get("url"),
                record.get("notes"),
            ),
        )
        return "inserted"


def import_record(db: sqlite3.Connection, record: dict) -> str:
    if record.get("id") is not None:
        return update_by_id(db, record)
    return upsert_by_callsign(db, record)


def move_to_backup(src: Path) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    dst = BACKUP_DIR / src.name

    if not dst.exists():
        shutil.move(str(src), str(dst))
        return dst

    stem = src.stem
    suffix = src.suffix
    counter = 1
    while True:
        candidate = BACKUP_DIR / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            shutil.move(str(src), str(candidate))
            return candidate
        counter += 1


def process_file(db: sqlite3.Connection, path: Path) -> tuple[int, int]:
    print(f"processing: {path}")
    records = load_records_from_xlsx(path)

    inserted = 0
    updated = 0

    for record in records:
        action = import_record(db, record)
        label = f"id={record['id']}" if record.get("id") is not None else (
            record.get("callsign") or "").strip().upper()
        print(f"{action}: {label}")
        if action == "inserted":
            inserted += 1
        else:
            updated += 1

    db.commit()
    backup_path = move_to_backup(path)
    print(f"moved to backup: {backup_path}")

    return inserted, updated


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(DATA_DIR.glob(FILE_GLOB))
    if not files:
        print(f"no files matched: {DATA_DIR / FILE_GLOB}", file=sys.stderr)
        return

    db = sqlite3.connect(str(DB_PATH))
    try:
        ensure_schema(db)

        total_inserted = 0
        total_updated = 0
        failed = 0

        for path in files:
            try:
                inserted, updated = process_file(db, path)
                total_inserted += inserted
                total_updated += updated
            except Exception as e:
                db.rollback()
                print(f"Error processing {path}: {e}", file=sys.stderr)
                failed += 1

        print(
            f"done: inserted={total_inserted}, updated={total_updated}, failed_files={failed}")

        if failed:
            sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
