#!/usr/bin/env python3
import sys
import json
from datetime import datetime
from openpyxl import load_workbook


def normalize(value):
    """Convert Excel values into JSON-safe values."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value


def main():
    if len(sys.argv) != 2:
        print("Usage: excel2json.py file.xlsx", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("[]")
        return

    header = rows[0]

    records = []
    for row in rows[1:]:
        record = {}
        for key, value in zip(header, row):
            record[key] = normalize(value)
        records.append(record)

    json.dump(records, sys.stdout, ensure_ascii=False, indent=2)
    print()


if __name__ == "__main__":
    main()
